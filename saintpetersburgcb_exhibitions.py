# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# radio
import re
from datetime import datetime, date

import dateparser
import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
SEARCH_PAGE_URL = "https://saintpetersburgcb.com/exhibitions/?page=%s"
PAGE_URL = "https://saintpetersburgcb.com"
DATA_CARD = []
category = {
    # "Общая выставочная площадь, кв.м.": "G%s",
    # "Общая вместимость, чел.": "H%s",
    # "Транспортная доступность": "I%s",
    # "Парковка": "J%s",
    # "Дополнительные возможности": "K%s",
    # "Удаленность от аэропорта, км.": "L%s",
    # "Удаленность от Московского вокзала, км.": "M%s",
    # "Общее количество залов": "N%s",
}


def parsing_exhibitions():
    wb = load_workbook(filename='Exhibitions.xlsx')
    sheet = wb['Лист1']
    ws = wb.active

    articles = []
    page = 1
    row = 2
    hrefs = []
    while page < 1000:
        body = []
        is_not_stopped, body, hrefs = get_urls(body, page, hrefs=hrefs)
        page += 1
        if not is_not_stopped:
            break
        print(page)
        i = 1

        for article in body:
            try:
                is_time, articles, res = get_page(articles, article)
                if res is not None:
                    ws["A%s" % row] = res.get("href")
                    ws["B%s" % row] = res.get("title")
                    ws["C%s" % row] = res.get("text")
                    ws["D%s" % row] = res.get("img")
                    ws["E%s" % row] = "\n ".join(res.get("text_images"))
                    ws["F%s" % row] = res.get("addr")
                    ws["G%s" % row] = res.get("site")


                    row += 1
                    i += 1

            except Exception:
                pass
        if i != 19:
            print("ERROR ")

    wb.save(f"Exhibitions{page}.xlsx")
    return articles


def get_urls(body, page, hrefs, attempts=0):
    try:
        res = requests.get(SEARCH_PAGE_URL % page,
                           headers={
                               "user-agent": USER_AGENT
                           },

                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
    except Exception as e:
        # logger.info(str(e))
        if attempts < 10:
            return get_urls(page, hrefs, attempts + 1)
        return False, body, hrefs
    if res.ok:
        soup = BeautifulSoup(res.text)

        articles = soup.find("div", {"class": "row events"}).find_all("div", {"class": "events-container"})

        if len(articles) == 0:
            return False, body, hrefs

        for article in articles:
            try:
                href = PAGE_URL + article.find("a").attrs["href"]
                if href in hrefs:
                    return False, body, hrefs
                hrefs.append(href)
                body.append(
                    {
                        "href": href,
                        "img": PAGE_URL + article.find("img").attrs["src"],
                        "title": article.find("span", {"class":"text"}).text.strip(),
                    }
                )
            except Exception:
                pass

    elif res.status_code == 404:
        return False, body, hrefs
    return True, body, hrefs


def get_page(articles, article_body, attempt=0):
    try:
        url = article_body['href']

        res = requests.get(url, headers={
            "user-agent": USER_AGENT
        },
                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
        if res.ok:
            soup_all = BeautifulSoup(res.text)
            soup = soup_all.find("div", {"class": "column-14 text-align-justify"})

            text_images = []
            try:
                for i in soup_all.find("div", {"class": "fotorama"}).find_all("a"):
                    text_images.append(PAGE_URL + i.attrs['href'])
            except Exception:
                pass
            try:
                for i in soup.find_all("img"):
                    text_images.append(PAGE_URL + i.attrs['src'])
            except Exception:
                pass
            text = ""
            for p in soup.find_all("p"):
                text += p.text.strip() + "\n"
            site = ""
            addr = ""
            for d in soup_all.find("div", {"class": "row item-info"}).find_all("div", {"class": "row"}):
                try:
                    if 'Место проведения' in d.find("div", {"class":"column-5"}):
                        addr = d.find("div", {"class":"column-8 offset-1"}).text.strip()
                    if 'Сайт' in d.find("div", {"class": "column-5"}):
                        site = d.find("div", {"class": "column-8 offset-1"}).text.strip()
                except Exception:
                    pass
            res = {"href": article_body['href'],
                   "title": article_body['title'],
                   "img": article_body['img'],
                   "text": text,
                   "text_images": text_images,
                   "site": site,
                   "addr": addr,
                   }

            articles.append(
                res)

            return True, articles, res
        return False, articles, None
    except Exception as e:
        if attempt > 2:
            return False, articles, None
        return get_page(articles, article_body, attempt + 1)


def update_proxy(proxy):
    return None


DEFAULTS_TIMEOUT = 100

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    articles = parsing_exhibitions()
