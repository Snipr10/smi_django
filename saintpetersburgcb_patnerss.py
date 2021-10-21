# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# radio
import re
from datetime import datetime, date

import dateparser
import requests
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import load_workbook

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
SEARCH_PAGE_URL = "https://saintpetersburgcb.com/partners/?page=%s"
PAGE_URL = "https://saintpetersburgcb.com"
CONTACT = []


def parsing_partners():
    wb = load_workbook(filename='Partners.xlsx')
    sheet = wb['Лист1']
    ws = wb.active

    articles = []
    page = 1
    row = 2
    hrefs = []
    while page < 1000:
        body = []
        is_not_stopped, body, hrefs = get_urls(body, page, hrefs=hrefs)
        page += 1
        if not is_not_stopped:
            break
        print(page)
        i = 1

        for article in body:
            try:
                is_time, articles, res = get_page(articles, article)
                if res is not None:
                    "ссылка	Картинка ( url)	Название	Текст	Сайт	Адрес	Телефон	Электронная почта"
                    ws["A%s" % row] = res.get("href")
                    ws["B%s" % row] = res.get("img")
                    ws["C%s" % row] = res.get("title")
                    ws["D%s" % row] = res.get("text")
                    ws["E%s" % row] = res.get("site")
                    ws["F%s" % row] = res.get("addr")
                    ws["G%s" % row] = res.get("phone")
                    ws["H%s" % row] = res.get("email")
                    row += 1
                    i += 1

            except Exception:
                pass

        if i != 13:
            print("ERROR ")
        if page % 50 == 0:
            wb.save(f"Partners{page}.xlsx")

    wb.save(f"Partners{page}.xlsx")
    return articles


def get_urls(body, page, hrefs, attempts=0):
    try:
        res = requests.get(SEARCH_PAGE_URL % page,
                           headers={
                               "user-agent": USER_AGENT
                           },

                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
    except Exception as e:
        # logger.info(str(e))
        if attempts < 10:
            return get_urls(page, hrefs, attempts + 1)
        return False, body, hrefs
    if res.ok:
        soup = BeautifulSoup(res.text)

        articles = soup.find("div", {"class": "tiles"}).find_all("div")

        if len(articles) == 0:
            return False, body, hrefs
        for article in articles:
            try:
                article_text = article.find("a")
                href = PAGE_URL + article_text.attrs.get("href")
                if href in hrefs:
                    return False, body, hrefs
                hrefs.append(href)

                body.append(
                    {
                        "href": href,
                        "img": PAGE_URL + article.find("img").attrs["src"],
                        "title": article_text.text,
                    }
                )
            except Exception:
                pass

    elif res.status_code == 404:
        return False, body, hrefs
    return True, body, hrefs


def get_page(articles, article_body, attempt=0):
    try:
        url = article_body['href']

        res = requests.get(url, headers={
            "user-agent": USER_AGENT
        },
                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
        if res.ok:
            soup_all = BeautifulSoup(res.text)
            soup_text = soup_all.find("div", {"class": "description text-align-justify"})
            text = ""
            for content in soup_text.find("p").contents:
                if isinstance(content, Tag):
                    text += "\n"
                else:
                    text += str(content)
            soup_contanct = soup_all.find("dl", {"class":"list-of-params"})
            contact_text =  soup_contanct.find_all("dt")
            contact_data =  soup_contanct.find_all("dd")
            site = ""
            addr = ""
            phone = ""
            email = ""
            for i in range(len(contact_text)):
                if contact_text[i].text == "Сайт":
                    site = contact_data[i].text
                elif contact_text[i].text == "Адрес":
                    addr = contact_data[i].text
                elif contact_text[i].text == 'Телефон':
                    phone = contact_data[i].text
                elif contact_text[i].text == 'Эл. почта':
                    email = contact_data[i].text
            for dt in soup_contanct.find_all("dt"):
                if dt.text not in CONTACT:
                    CONTACT.append(dt.text)
            res = {"href": article_body['href'],
                   "img": article_body['img'],
                   "title": article_body['title'],
                   "text": text.strip(),
                   "site": site,
                   "addr": addr,
                   "phone": phone,
                   "email": email,
                   }
            articles.append(
                res)

            return True, articles, res
        return False, articles, None
    except Exception as e:
        if attempt > 2:
            return False, articles, None
        return get_page(articles, article_body, attempt + 1)


def update_proxy(proxy):
    return None


DEFAULTS_TIMEOUT = 100

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    articles = parsing_partners()
