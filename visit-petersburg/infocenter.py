# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# radio
import json
import re
from datetime import datetime, date

import dateparser
import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
SEARCH_PAGE_URL = "https://www.visit-petersburg.ru/ru/infocenter/?name=&page=%s"
DATA_CARD = []
category = {
    # "Общая выставочная площадь, кв.м.": "G%s",
    # "Общая вместимость, чел.": "H%s",
    # "Транспортная доступность": "I%s",
    # "Парковка": "J%s",
    # "Дополнительные возможности": "K%s",
    # "Удаленность от аэропорта, км.": "L%s",
    # "Удаленность от Московского вокзала, км.": "M%s",
    # "Общее количество залов": "N%s",
}


def parsing_infocenter():
    wb = load_workbook(filename='infocenter.xlsx')
    sheet = wb['Лист1']
    ws = wb.active

    articles = []
    page = 1
    row = 2
    hrefs = []
    while page < 1000:
        body = []
        is_not_stopped, body, = get_urls(body, page)
        page += 1
        if not is_not_stopped:
            break
        print(page)
        i = 1

        for article in body:
            try:
                if article is not None:
                    ws["A%s" % row] = article.get("id")
                    ws["B%s" % row] = "https://www.visit-petersburg.ru" + article.get("photo")
                    ws["C%s" % row] = article.get("name")
                    ws["D%s" % row] = article.get("address")
                    ws["E%s" % row] = article.get("description")
                    ws["F%s" % row] = article.get("tourist_help")
                    ws["G%s" % row] = article.get("tickets")
                    ws["H%s" % row] = article.get("souvenirs")
                    ws["I%s" % row] = article.get("history")
                    ws["J%s" % row] = article.get("phone")
                    ws["K%s" % row] = article.get("fax")
                    ws["L%s" % row] = article.get("domain")
                    ws["M%s" % row] = article.get("email")
                    ws["N%s" % row] = article.get("work_time")


                    row += 1
                    i += 1
                else:
                    print(1)

            except Exception:
                pass
    wb.save(f"infocenter{page}.xlsx")
    return articles


def get_urls(body, page, attempts=0):
    try:
        res = requests.get(SEARCH_PAGE_URL % page,
                           headers={
                               "user-agent": USER_AGENT
                           },

                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
    except Exception as e:
        # logger.info(str(e))
        if attempts < 10:
            return get_urls(page, attempts + 1)
        return False, body
    if res.ok:
        soup = BeautifulSoup(res.text)

        articles = json.loads(soup.find("div", {"class": "consulate"}).attrs["data-widget"])['react']['props']['consulates']

        if len(articles) == 0:
            return False, body

        body.extend(articles)
    elif res.status_code == 404:
        return False, body
    return True, body



DEFAULTS_TIMEOUT = 100

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    articles = parsing_infocenter()
