# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# radio
import re
from datetime import datetime, date

import dateparser
import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
SEARCH_PAGE_URL = "https://www.visit-petersburg.ru/ru/news/?page=%s"
PAGE_URL = "https://www.visit-petersburg.ru"
DATA_CARD = []
category = {
    # "Общая выставочная площадь, кв.м.": "G%s",
    # "Общая вместимость, чел.": "H%s",
    # "Транспортная доступность": "I%s",
    # "Парковка": "J%s",
    # "Дополнительные возможности": "K%s",
    # "Удаленность от аэропорта, км.": "L%s",
    # "Удаленность от Московского вокзала, км.": "M%s",
    # "Общее количество залов": "N%s",
}


def parsing_news():
    wb = load_workbook(filename='news.xlsx')
    sheet = wb['Лист1']
    ws = wb.active

    articles = []
    page = 1
    row = 2
    hrefs = []
    while page < 2000:
        body = []
        is_not_stopped, body, hrefs = get_urls(body, page, hrefs=hrefs)
        page += 1
        if not is_not_stopped:
            break
        print(page)
        i = 1

        for article in body:
            try:

                is_time, articles, res = get_page(articles, article)
                article["href"] = article["href"].replace("https://www.visit-petersburg.ru/ru/",
                                                          'https://www.visit-petersburg.ru/en/')
                is_time, articles, res_eng = get_page(articles, article)

                if res is not None:
                    ws["A%s" % row] = res.get("href")
                    ws["B%s" % row] = res.get("title")
                    ws["C%s" % row] = dateparser.parse(res.get("date")).date()
                    ws["D%s" % row] = res.get("text")
                    ws["E%s" % row] = "\n ".join(res.get("images"))
                    ws["F%s" % row] = res.get("type")
                    ws["G%s" % row] = res.get("prev_photo")

                    if res_eng is not None:
                        ws["H%s" % row] = res_eng.get("title")
                        ws["I%s" % row] = dateparser.parse(res_eng.get("date")).date()
                        ws["J%s" % row] = res_eng.get("text")
                        ws["K%s" % row] = res_eng.get("type")

                    row += 1
                    i += 1
                else:
                    print(1)
            except Exception:
                pass

    wb.save(f"news{page}.xlsx")
    return articles


def get_urls(body, page, hrefs, attempts=0):
    try:

        res = requests.get(SEARCH_PAGE_URL % page,
                           headers={
                               'X-Requested-With': 'XMLHttpRequest',
                                "user-agent": USER_AGENT,
                           },

                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
    except Exception as e:
        # logger.info(str(e))
        if attempts < 10:
            return get_urls(page, hrefs, attempts + 1)
        return False, body, hrefs
    if res.ok:
        soup = BeautifulSoup(res.json()["news"])

        articles = soup.find_all("div", {"class": "img_wrapper"})

        if len(articles) == 0:
            return False, body, hrefs

        for article in articles:
            try:
                href = PAGE_URL + article.find("a").attrs["href"]
                if href in hrefs:
                    return False, body, hrefs
                hrefs.append(href)

                body.append(
                    {
                        "href": href,
                        "prev_photo": PAGE_URL + article.find("img").attrs.get("src")
                    }
                )

            except Exception:
                pass

    elif res.status_code == 404:
        return False, body, hrefs
    return True, body, hrefs


def get_page(articles, article_body, attempt=0):
    try:
        url = article_body['href']

        res = requests.get(url, headers={
            "user-agent": USER_AGENT
        },
                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
        if res.ok:
            soup_all = BeautifulSoup(res.text)
            text = soup_all.find("h3").text.strip()
            if text == 'See also':
                text = ""
            for p in soup_all.find("div", {"class": "text_content"}).find_all("p"):
                text += "\n" + p.text.strip()
            images = []
            for i in soup_all.find("div", {"class": "carousel-inner"}).find_all("div", {"class": "item active"}):
                images.append(PAGE_URL + i.find("img").attrs["src"])
            res = {"href": article_body['href'],
                   "prev_photo": article_body['prev_photo'],
                   "title": soup_all.find("h1").text.strip(),
                   "date": soup_all.find("div", {"class": "news_date"}).text.strip(),
                   "text": text,
                   "images": images,
                   "type": soup_all.find("div", {"class": "bread_crumbs"}).text.strip(),
                   }

            articles.append(
                res)

            return True, articles, res
        return False, articles, None
    except Exception as e:
        if attempt > 2:
            return False, articles, None
        return get_page(articles, article_body, attempt + 1)


def update_proxy(proxy):
    return None


DEFAULTS_TIMEOUT = 100

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    articles = parsing_news()
