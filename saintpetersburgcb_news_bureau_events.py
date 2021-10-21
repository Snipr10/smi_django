# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# radio
import re
from datetime import datetime, date

import dateparser
import requests
from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
SEARCH_PAGE_URL = "https://saintpetersburgcb.com/news/bureau_events/?page=%s"
PAGE_URL = "https://saintpetersburgcb.com"

category = {"Амбассадоры": "H%s", "Бюро":"I%s", "Комитет":"J%s", "Партнёры":"L%s", "Устойчивое развитие":"L%s"}

def parsing_news():
    wb = load_workbook(filename='News.xlsx')
    sheet = wb['Лист1']
    ws = wb.active

    articles = []
    page = 1
    row = 2
    hrefs = []
    while page < 1000:
        body = []
        is_not_stopped, body, hrefs = get_urls(body, page, hrefs=hrefs)
        page += 1
        if not is_not_stopped:
            break
        print(page)
        i = 1

        for article in body:
            try:
                is_time, articles, res = get_page(articles, article)
                if res is not None:
                    ws["A%s" % row] = res.get("href")
                    ws["B%s" % row] = res.get("title")
                    ws["C%s" % row] = res.get("anons")
                    ws["D%s" % row] = res.get("date").date()
                    ws["E%s" % row] = res.get("img")
                    ws["F%s" % row] = res.get("text")
                    ws["G%s" % row] = "\n ".join(res.get("text_images"))

                    if res.get("news_category") is not None:
                        for c in res.get("news_category").find_all("a"):
                            try:
                                ws[category.get(c.text) % row] = c.text
                            except Exception:
                                print(c)
                    row += 1
                    i += 1

            except Exception:
                pass

        if i != 6:
            print("ERROR ")
        if page % 50 == 0:
            wb.save(f"News_b{page}.xlsx")

    wb.save(f"News_b{page}.xlsx")
    return articles


def get_urls(body, page, hrefs, attempts=0):
    try:
        res = requests.get(SEARCH_PAGE_URL % page,
                           headers={
                               "user-agent": USER_AGENT
                           },

                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
    except Exception as e:
        # logger.info(str(e))
        if attempts < 10:
            return get_urls(page, hrefs, attempts + 1)
        return False, body, hrefs
    if res.ok:
        soup = BeautifulSoup(res.text)

        articles = soup.find("div", {"class": "column-15 offset-2"}).find_all("div", {"class": "row"})

        if len(articles) == 0:
            return False, body, hrefs
        for article in articles:
            try:
                article_event = article.find("a", {"class": "event-image"})
                article_text = article.find("div", {"class": "column-9 offset-small"})
                href = PAGE_URL + article_event.attrs.get("href")
                if href in hrefs:
                    return False, body, hrefs
                hrefs.append(href)
                try:
                    news_category = article.find("div", {"class": "news-category"})

                except Exception:
                    news_category = None

                body.append(
                    {
                        "href": PAGE_URL + article_event.attrs.get("href"),
                        "img": PAGE_URL + article_event.find("img").attrs.get("src"),
                        "title": article_text.find("a").text,
                        "text": article_text.find("p").text,
                        "news_category": news_category
                    }
                )
            except Exception:
                pass

    elif res.status_code == 404:
        return False, body, hrefs
    return True, body, hrefs


def get_page(articles, article_body, attempt=0):
    try:
        url = article_body['href']

        res = requests.get(url, headers={
            "user-agent": USER_AGENT
        },
                           # proxies=proxy.get(list(proxy.keys())[0]),
                           # timeout=DEFAULTS_TIMEOUT
                           )
        if res.ok:
            soup_all = BeautifulSoup(res.text)
            soup = soup_all.find("div", {"class": "column-15"})

            article_date = dateparser.parse(soup_all.find("div", {"class": "date"}).text.strip())
            text_images = []
            try:
                for i in soup_all.find("div", {"class": "fotorama"}).find_all("a"):
                    text_images.append(PAGE_URL + i.attrs['href'])
            except Exception:
                pass
            try:
                for i in soup.find_all("img"):
                    text_images.append(PAGE_URL + i.attrs['src'])
            except Exception:
                pass
            res = {"href": article_body['href'],
                   "title": article_body['title'],
                   "anons": article_body['text'],
                   "date": article_date,
                   "text": soup.text.strip(),
                   "img": article_body['img'],
                   "news_category": article_body['news_category'],
                   "text_images": text_images
                   }
            articles.append(
                res)

            return True, articles, res
        return False, articles, None
    except Exception as e:
        if attempt > 2:
            return False, articles, None
        return get_page(articles, article_body, attempt + 1)


def update_proxy(proxy):
    return None


DEFAULTS_TIMEOUT = 100

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    articles = parsing_news()
